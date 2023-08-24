import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.formatting.rule import FormulaRule
import pandas as pd
import datetime


class xlreport(object):
    """The class is used to generate the excel report for programming voltage and
    current accuracy tests.


    """

    def __init__(self):
        """ "Initialize certain parameter for the excel sheet such as font, colour
        fill, the path where the excel path is also generated here.

        Excel files are generated in Excel files Folder in this repository.


        """
        self.red_font = Font(size=14, bold=True, color="ffffff")
        self.red_fill = PatternFill(
            start_color="ffcccc", end_color="ffcccc", fill_type="solid"
        )
        self.green_font = Font(size=14, bold=True, color="013220")
        self.green_fill = PatternFill(
            start_color="FFAAFF00", end_color="FFAAFF00", fill_type="solid"
        )
        self.path = (
            r"C:\Users\zhiywong\OneDrive - Keysight Technologies\Documents\GitHub\PyVisa\Excel Files\\"
            + datetime.datetime.now().strftime("%Y-%m-%d--%H-%M-%S")
            + ".xlsx"
        )

    def adjustcolumnWidth(self, worksheet, value):
        """To adjust the column width from column A to I

        Args:
            worksheet: string specifying which worksheet the column width to be changed.
            value: integer specifiying the width value to be changed to.
        """
        for x in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
            worksheet.column_dimensions[x].width = value

    def alignCell(self, worksheet, row_number, column_number, style):
        """To change the alignment in cells

        Args:
            worksheet: string specifying the worksheet where the cells will be aligned.
            row_number: integer specifying the row in the worksheet the cell to be aligned is in.
            column_number: integer specifying the column in the worksheet the cell to be aligned is in.
            style: string specifying the type of alignment the cell should be in.
        """
        currentCell = worksheet.cell(row=row_number, column=column_number)
        currentCell.alignment = Alignment(horizontal=style)

    def run(self):
        """Function to execute the generate of the excel sheet

        The function begins by improrting dataframes from csv files generated from
        data.py. These dataframes will be labeled and placed at a specific position in the
        excel report. The graph generated will also be imported into the excel report


        """
        with pd.ExcelWriter(self.path, engine="openpyxl") as writer:
            df1 = pd.read_csv("csv/error.csv", index_col=False)
            df2 = pd.read_csv(
                "csv/instrumentData.csv",
                index_col=False,
            )

            df4 = pd.read_csv("csv/config.csv")
            df1.to_excel(writer, sheet_name="Data", index=False, startrow=7, startcol=3)
            df2.to_excel(writer, sheet_name="Data", index=False)
            df4.to_excel(writer, sheet_name="Data", index=False, startrow=6)
            wb = writer.book
            ws = wb["Data"]
            cellref = "N9:N" + str(ws.max_row)

            # Conditional Formatting to set Font and Colour of Cell depending on boolList
            ws.conditional_formatting.add(
                cellref,
                FormulaRule(
                    formula=[f'NOT(ISERROR(SEARCH("PASS",{cellref})))'],
                    stopIfTrue=True,
                    fill=self.green_fill,
                    font=self.green_font,
                ),
            )
            ws.conditional_formatting.add(
                cellref,
                FormulaRule(
                    formula=[f'NOT(ISERROR(SEARCH("FAIL",{cellref})))'],
                    stopIfTrue=True,
                    fill=self.red_fill,
                    font=self.red_font,
                ),
            )

            self.adjustcolumnWidth(ws, 20)
            self.adjustcolumnWidth(ws, 20)
            self.alignCell(ws, 1, 1, "left")
            self.alignCell(ws, 2, 1, "left")
            self.alignCell(ws, 3, 1, "left")
            self.alignCell(ws, 4, 1, "left")

            ws.cell(row=7, column=4).value = "Time Generated: "
            ws.cell(row=7, column=5).value = datetime.datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            )

            # Inserting graph of test into excel report
            img = openpyxl.drawing.image.Image("images/Chart.png")
            img.anchor = "R1"
            ws.add_image(img)

            wb.save(self.path)
